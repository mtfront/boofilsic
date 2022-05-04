import openpyxl
import requests
import re
from lxml import html
from markdownify import markdownify as md
from datetime import datetime
from common.scraper import get_scraper_by_url
import logging
import pytz
from django.conf import settings
from django.core.exceptions import ObjectDoesNotExist
from user_messages import api as msg
import django_rq
from common.utils import GenerateDateUUIDMediaFilePath
import os
from books.models import BookReview, Book
from movies.models import MovieReview, Movie
from music.models import AlbumReview, Album
from games.models import GameReview, Game
from common.scraper import DoubanAlbumScraper, DoubanBookScraper, DoubanGameScraper, DoubanMovieScraper
from PIL import Image
from io import BytesIO
import filetype


logger = logging.getLogger(__name__)


def fetch_remote_image(url):
    try:
        print(f'fetching remote image {url}')
        raw_img = None
        ext = None
        if settings.SCRAPESTACK_KEY is not None:
            dl_url = f'http://api.scrapestack.com/scrape?access_key={settings.SCRAPESTACK_KEY}&url={url}'
        elif settings.SCRAPERAPI_KEY is not None:
            dl_url = f'http://api.scraperapi.com?api_key={settings.SCRAPERAPI_KEY}&url={url}'
        else:
            dl_url = url
        img_response = requests.get(dl_url, timeout=settings.SCRAPING_TIMEOUT)
        raw_img = img_response.content
        img = Image.open(BytesIO(raw_img))
        img.load()  # corrupted image will trigger exception
        content_type = img_response.headers.get('Content-Type')
        ext = filetype.get_type(mime=content_type.partition(';')[0].strip()).extension
        f = GenerateDateUUIDMediaFilePath(None, "x." + ext, settings.MARKDOWNX_MEDIA_PATH)
        file = settings.MEDIA_ROOT + f
        local_url = settings.MEDIA_URL + f
        os.makedirs(os.path.dirname(file), exist_ok=True)
        img.save(file)
        print(f'remote image saved as {local_url}')
        return local_url
    except Exception:
        print(f'unable to fetch remote image {url}')
        return url


class DoubanImporter:
    total = 0
    skipped = 0
    imported = 0
    failed = []
    user = None
    visibility = 0
    file = None

    def __init__(self, user, visibility):
        self.user = user
        self.visibility = visibility

    def update_user_import_status(self, status):
        self.user.preference.import_status['douban_pending'] = status
        self.user.preference.import_status['douban_file'] = self.file
        self.user.preference.import_status['douban_visibility'] = self.visibility
        self.user.preference.import_status['douban_total'] = self.total
        self.user.preference.import_status['douban_skipped'] = self.skipped
        self.user.preference.import_status['douban_imported'] = self.imported
        self.user.preference.import_status['douban_failed'] = self.failed
        self.user.preference.save(update_fields=['import_status'])

    def import_from_file(self, uploaded_file):
        try:
            wb = openpyxl.open(uploaded_file, read_only=True, data_only=True, keep_links=False)
            wb.close()
            file = settings.MEDIA_ROOT + GenerateDateUUIDMediaFilePath(None, "x.xlsx", settings.SYNC_FILE_PATH_ROOT)
            os.makedirs(os.path.dirname(file), exist_ok=True)
            with open(file, 'wb') as destination:
                for chunk in uploaded_file.chunks():
                    destination.write(chunk)
            self.file = file
            self.update_user_import_status(2)
            django_rq.get_queue('doufen').enqueue(self.import_from_file_task)
        except Exception:
            return False
        # self.import_from_file_task(file, user, visibility)
        return True

    def import_from_file_task(self):
        msg.info(self.user, f'开始导入豆瓣评论')
        self.update_user_import_status(1)
        f = open(self.file, 'rb')
        wb = openpyxl.load_workbook(f, read_only=True, data_only=True, keep_links=False)
        self.import_sheet(wb['书评'], DoubanBookScraper, Book, BookReview)
        self.import_sheet(wb['影评'], DoubanMovieScraper, Movie, MovieReview)
        self.import_sheet(wb['乐评'], DoubanAlbumScraper, Album, AlbumReview)
        self.import_sheet(wb['游戏评论&攻略'], DoubanGameScraper, Game, GameReview)
        self.update_user_import_status(0)
        msg.success(self.user, f'豆瓣评论导入完成，共处理{self.total}篇，已存在{self.skipped}篇，新增{self.imported}篇。')
        if len(self.failed):
            msg.error(self.user, f'豆瓣评论导入时未能处理以下网址：\n{" , ".join(self.failed)}')

    def import_sheet(self, worksheet, scraper, entity_class, review_class):
        prefix = f'{self.user} {review_class.__name__} |'
        if worksheet is None:  # or worksheet.max_row < 2:
            print(f'{prefix} empty sheet')
            return
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            cells = [cell for cell in row]
            if len(cells) < 6:
                continue
            title = cells[0]
            review_url = cells[2]
            time = cells[3]
            content = cells[6]
            self.total += 1
            if time:
                time = datetime.strptime(time, "%Y-%m-%d %H:%M:%S")
                tz = pytz.timezone('Asia/Shanghai')
                time = time.replace(tzinfo=tz)
            else:
                time = None
            if not content:
                content = ""
            if not title:
                title = ""
            r = self.import_review(title, review_url, content, time, scraper, entity_class, review_class)
            if r == 1:
                self.imported += 1
            elif r == 2:
                self.skipped += 1
            else:
                self.failed.append(review_url)
            self.update_user_import_status(1)

    def import_review(self, title, review_url, content, time, scraper, entity_class, review_class):
        # return 1: done / 2: skipped / None: failed
        prefix = f'{self.user} {review_class.__name__} |'
        url = None
        print(f'{prefix} fetching {review_url}')
        try:
            if settings.SCRAPESTACK_KEY is not None:
                _review_url = f'http://api.scrapestack.com/scrape?access_key={settings.SCRAPESTACK_KEY}&url={review_url}'
            else:
                _review_url = review_url
            r = requests.get(_review_url, timeout=settings.SCRAPING_TIMEOUT)
            if r.status_code != 200:
                print(f'{prefix} fetching error {review_url} {r.status_code}')
                return
            h = html.fromstring(r.content.decode('utf-8'))
            for u in h.xpath("//header[@class='main-hd']/a/@href"):
                if '.douban.com/subject/' in u:
                    url = u
            if not url:
                print(f'{prefix} fetching error {review_url} unable to locate url')
                return
        except Exception:
            print(f'{prefix} fetching exception {review_url}')
            return
        try:
            entity = entity_class.objects.get(source_url=url)
            print(f'{prefix} matched {url}')
        except ObjectDoesNotExist:
            try:
                print(f'{prefix} scraping {url}')
                scraper.scrape(url)
                form = scraper.save(request_user=self.user)
                entity = form.instance
            except Exception as e:
                print(f"{prefix} scrape failed: {url} {e}")
                logger.error(f"{prefix} scrape failed: {url}", exc_info=e)
                return
        params = {
            'owner': self.user,
            entity_class.__name__.lower(): entity
        }
        if review_class.objects.filter(**params).exists():
            return 2
        content = re.sub(r'<span style="font-weight: bold;">([^<]+)</span>', r'<b>\1</b>', content)
        content = re.sub(r'<div class="image-caption">([^<]+)</div>', r'<br><i>\1</i><br>', content)
        content = md(content)
        content = re.sub(r'(?<=!\[\]\()([^)]+)(?=\))', lambda x: fetch_remote_image(x[1]), content)
        params = {
            'owner': self.user,
            'created_time': time,
            'edited_time': time,
            'title': title,
            'content': content,
            'visibility': self.visibility,
            entity_class.__name__.lower(): entity,
        }
        review_class.objects.create(**params)
        return 1
